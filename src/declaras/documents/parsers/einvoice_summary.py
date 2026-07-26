"""Lector del resumen de facturas electronicas que entrega el portal en XLSX.

El archivo tiene un bloque de encabezado clave-valor (contribuyente y periodo) y luego
una tabla con una factura por fila. La columna "Valor Susceptible Beneficio" ya viene
filtrada por la DIAN segun el medio de pago: si la factura se pago en efectivo, ese valor
llega en cero, porque la deduccion del 1% exige pago por canal financiero. Por eso el
total de esa columna es directamente la base de la deduccion, sin logica adicional.
"""

from __future__ import annotations

import hashlib
import re
from datetime import datetime, time
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from declaras.documents.models import (
    Confidence,
    DocumentReading,
    ExtractedField,
    ExtractedRow,
    ReadingWarning,
)
from declaras.domain.errors import DocumentUnreadableError
from declaras.observability import get_logger

log = get_logger(__name__)

PARSER_NAME = "einvoice_summary.xlsx.v1"

# Etiquetas del bloque de encabezado (columna A) y el nombre logico de cada una.
_HEADER_LABELS = {
    "Año Gravable": "tax_year",
    "Doc Identificacion Adq.": "id_number",
    "Nombre o razón social": "taxpayer_name",
}

# Encabezado de la tabla de facturas: se busca por texto, no por fila fija.
_TABLE_HEADER_MARKER = "Identificación Emisor Factura"

_COL_ISSUER_NIT = 1
_COL_ISSUER_NAME = 2
_COL_ISSUE_DATE = 3
_COL_INVOICED_AMOUNT = 4
_COL_CREDIT_NOTES = 5
_COL_DEBIT_NOTES = 6
_COL_NET_AMOUNT = 7
_COL_BENEFIT_ELIGIBLE_AMOUNT = 8
_COL_PAYMENT_METHOD = 9
_COL_INVOICE_NUMBER = 10
_COL_CUFE = 11

_CASH_PAYMENT_METHOD = "efectivo"


def parse(content: bytes) -> DocumentReading:
    """Lee el resumen de facturas electronicas y calcula la base de la deduccion del 1%."""
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise DocumentUnreadableError(
            "el archivo no es un XLSX legible", parser=PARSER_NAME
        ) from exc

    sheet = workbook.active
    if sheet is None:
        raise DocumentUnreadableError("el XLSX no tiene hojas", parser=PARSER_NAME)

    warnings: list[ReadingWarning] = []
    fields = _read_header(sheet)
    rows, header_row = _read_invoices(sheet, warnings)
    fields += _summary_fields(rows, header_row)

    log.info("documents.einvoice_summary.parsed", invoices=len(rows), warnings=len(warnings))
    return DocumentReading(
        doc_type="EINVOICE_SUMMARY",
        parser=PARSER_NAME,
        content_sha256=hashlib.sha256(content).hexdigest(),
        fields=fields,
        rows=rows,
        warnings=warnings,
    )


def _read_header(sheet: Worksheet) -> list[ExtractedField]:
    """Bloque clave-valor de las primeras filas: se busca por la etiqueta en columna A."""
    fields: list[ExtractedField] = []
    for row in range(1, 10):
        label = _clean_text(sheet.cell(row=row, column=1).value)
        name = _HEADER_LABELS.get(label or "")
        if name is None:
            continue
        raw = sheet.cell(row=row, column=2).value
        value: Any = _as_int(raw) if name == "tax_year" else _clean_text(raw)
        fields.append(ExtractedField(name=name, value=value, source=f"B{row}"))
    return fields


def _read_invoices(
    sheet: Worksheet, warnings: list[ReadingWarning]
) -> tuple[list[ExtractedRow], int | None]:
    header_row = _find_table_header_row(sheet)
    if header_row is None:
        warnings.append(
            ReadingWarning(
                code="TABLE_HEADER_NOT_FOUND",
                message=(
                    "El resumen de facturas no tiene la forma esperada, así que no se pudo "
                    "leer el detalle. Hay que volver a traerlo del portal."
                ),
            )
        )
        return [], None

    rows: list[ExtractedRow] = []
    for row in range(header_row + 1, sheet.max_row + 1):
        issuer_nit = _clean_text(sheet.cell(row=row, column=_COL_ISSUER_NIT).value)
        amount = _as_int(sheet.cell(row=row, column=_COL_NET_AMOUNT).value)
        if not issuer_nit or amount is None:
            continue  # fila de cierre ("N facturas procesadas...") o vacia

        payment_method = _clean_text(sheet.cell(row=row, column=_COL_PAYMENT_METHOD).value)
        benefit_amount = (
            _as_int(sheet.cell(row=row, column=_COL_BENEFIT_ELIGIBLE_AMOUNT).value) or 0
        )
        is_cash = payment_method is not None and _CASH_PAYMENT_METHOD in payment_method.lower()
        if is_cash and benefit_amount:
            warnings.append(
                ReadingWarning(
                    code="CASH_PAYMENT_WITH_BENEFIT",
                    message=(
                        "Una factura pagada en efectivo aparece marcada como válida para el "
                        "descuento del 1%. En efectivo no aplica, así que hay que revisarla."
                    ),
                    source=f"fila {row}",
                )
            )

        rows.append(
            ExtractedRow(
                source=f"fila {row}",
                values={
                    "issuer_nit": issuer_nit,
                    "issuer_name": _clean_text(sheet.cell(row=row, column=_COL_ISSUER_NAME).value),
                    "issue_date": _clean_text(sheet.cell(row=row, column=_COL_ISSUE_DATE).value),
                    "invoiced_amount": _as_int(
                        sheet.cell(row=row, column=_COL_INVOICED_AMOUNT).value
                    ),
                    "credit_notes": _as_int(sheet.cell(row=row, column=_COL_CREDIT_NOTES).value),
                    "debit_notes": _as_int(sheet.cell(row=row, column=_COL_DEBIT_NOTES).value),
                    "net_amount": amount,
                    "benefit_eligible_amount": benefit_amount,
                    "payment_method": payment_method,
                    "invoice_number": _clean_text(
                        sheet.cell(row=row, column=_COL_INVOICE_NUMBER).value
                    ),
                    "cufe": _clean_text(sheet.cell(row=row, column=_COL_CUFE).value),
                },
            )
        )
    return rows, header_row


def _summary_fields(rows: list[ExtractedRow], header_row: int | None) -> list[ExtractedField]:
    if header_row is None:
        return []
    total_net = sum(int(r.values["net_amount"]) for r in rows)
    # Es la base directa de la deduccion del 1%: la DIAN ya excluyo los pagos en efectivo.
    total_benefit_eligible = sum(int(r.values["benefit_eligible_amount"]) for r in rows)
    return [
        ExtractedField(name="invoice_count", value=len(rows), confidence=Confidence.DETERMINISTIC),
        ExtractedField(
            name="total_net_amount",
            value=total_net,
            confidence=Confidence.DETERMINISTIC,
            unit="COP",
        ),
        ExtractedField(
            name="total_benefit_eligible_amount",
            value=total_benefit_eligible,
            confidence=Confidence.DETERMINISTIC,
            unit="COP",
        ),
    ]


def _find_table_header_row(sheet: Worksheet) -> int | None:
    for row in range(1, sheet.max_row + 1):
        if _clean_text(sheet.cell(row=row, column=1).value) == _TABLE_HEADER_MARKER:
            return row
    return None


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == time.min else value.isoformat()
    text = str(value).strip()
    return re.sub(r"\s+", " ", text) or None


def _as_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return int(value)
    digits = re.sub(r"[^\d-]", "", str(value))
    return int(digits) if digits and digits != "-" else None
