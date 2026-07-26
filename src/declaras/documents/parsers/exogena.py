"""Lector del reporte de informacion exogena que entrega el portal en XLSX.

El archivo tiene tres bloques: un encabezado con el contribuyente y el periodo, los cinco
topes con que la DIAN determina la obligacion de declarar, y el detalle de lo que cada
tercero reporto.

El detalle trae dos columnas que valen oro y ahorran el trabajo mas duro del motor:

  Detalle                      incluye el codigo oficial del concepto, por ejemplo
                               "Otros ingresos (Concepto: 5016)"
  Uso declaracion Sugerida     dice a que renglon del formulario 210 va el valor y a que
                               tope cuenta, por ejemplo "Tope 1: Ingresos brutos | R32
                               Ingresos brutos"

Es decir: no hay que adivinar en que casilla va cada concepto reportado, porque la propia
DIAN lo indica.

El detalle trae ademas a quien le reporto el tercero (numero de identificacion y nombre), y
eso no siempre es el contribuyente: un tercero puede reportar a la cedula correcta con el
nombre de otra persona. Es un error frecuente y una de las preguntas que el producto tiene
que poder responder, porque ese valor entra a los topes de obligacion como si fuera suyo.
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from datetime import datetime, time
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from declaras.documents.models import (
    DocumentReading,
    ExtractedField,
    ExtractedRow,
    ReadingWarning,
    ThresholdCode,
)
from declaras.domain.errors import DocumentUnreadableError
from declaras.observability import get_logger

log = get_logger(__name__)

PARSER_NAME = "exogena.xlsx.v1"

# Posiciones del encabezado, verificadas contra el reporte real. El segundo valor es como
# se llama el campo en un aviso que lee una persona: los avisos no deben filtrar nombres
# internos como "taxpayer_name".
_HEADER_CELLS: dict[str, tuple[str, str]] = {
    "cutoff_date": ("C3", "la fecha de corte"),
    "tax_year": ("C4", "el año gravable"),
    "id_kind_label": ("C6", "el tipo de documento"),
    "id_number": ("C7", "el número de identificación"),
    "taxpayer_name": ("C8", "el nombre del contribuyente"),
    "report_date": ("H2", "la fecha del reporte"),
}

# Filas de los topes y su codigo, en el orden en que el portal los imprime.
_THRESHOLD_ROWS: dict[int, ThresholdCode] = {
    15: ThresholdCode.INGRESOS,
    16: ThresholdCode.PATRIMONIO,
    17: ThresholdCode.CONSUMO_TARJETA,
    18: ThresholdCode.MOVIMIENTOS,
    19: ThresholdCode.COMPRAS,
}

_FIRST_DETAIL_ROW = 20

_COL_REPORTER_NIT = 1
_COL_REPORTER_NAME = 2
_COL_REPORTED_ID = 3
_COL_REPORTED_NAME = 4
_COL_CONCEPT = 5
_COL_AMOUNT = 6
_COL_SUGGESTED_USE = 7
_COL_EXTRA = 8

_CONCEPT_CODE_RE = re.compile(r"\(Concepto:\s*(\d+)\)")
_FORM_LINE_RE = re.compile(r"\bR(\d{1,3})\b")
_THRESHOLD_LABEL_RE = re.compile(r"Tope\s*(\d)")

# La DIAN no siempre suma los valores de un tope: para algunos compara dos fuentes y toma la
# mayor, y lo dice en el texto de la columna "Uso declaracion Sugerida" ("toma el mayor
# valor", "selecciona el mayor"). Distinguir esas filas es lo que permite explicar de donde
# sale un tope sin inventar una formula: son alternativas, no sumandos.
_COMPARED_NOT_ADDED_RE = re.compile(r"\b(?:el|la)\s+mayor\b", re.IGNORECASE)
_REPLACEMENT_CHAR = "�"

# Minimo de palabras en comun para considerar que dos nombres son de la misma persona. Los
# terceros escriben el nombre en cualquier orden ("JUAN JOSE VALENCIA MORENO" y "VALENCIA
# MORENO JUAN JOSE" son el mismo), asi que se comparan como conjuntos de palabras y no como
# texto. Dos palabras en comun aguantan que una venga mal escrita o con acentos danados, sin
# confundir a dos personas distintas: un nombre ajeno no comparte ninguna.
_MIN_SHARED_NAME_WORDS = 2


def parse(content: bytes) -> DocumentReading:
    """Lee el reporte de exogena y devuelve sus valores con procedencia."""
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:
        raise DocumentUnreadableError(
            "el archivo no es un XLSX legible", parser=PARSER_NAME
        ) from exc

    sheet = workbook.active
    if sheet is None:
        raise DocumentUnreadableError("el XLSX no tiene hojas", parser=PARSER_NAME)

    warnings: list[ReadingWarning] = []
    fields = _read_header(sheet, warnings)
    fields += _read_thresholds(sheet)
    rows = _read_details(sheet, warnings)
    _check_reported_to_taxpayer(fields, rows, warnings)

    log.info(
        "documents.exogena.parsed",
        fields=len(fields),
        rows=len(rows),
        warnings=len(warnings),
    )
    return DocumentReading(
        doc_type="EXOGENA",
        parser=PARSER_NAME,
        content_sha256=hashlib.sha256(content).hexdigest(),
        fields=fields,
        rows=rows,
        warnings=warnings,
    )


# ─────────────────────────── bloques del reporte ───────────────────────────


def _read_header(sheet: Worksheet, warnings: list[ReadingWarning]) -> list[ExtractedField]:
    fields: list[ExtractedField] = []
    for name, (cell, label) in _HEADER_CELLS.items():
        raw = sheet[cell].value
        if raw is None:
            warnings.append(
                ReadingWarning(
                    code="HEADER_FIELD_MISSING",
                    message=f"El reporte de la DIAN no trae {label}",
                    source=cell,
                )
            )
            continue
        value: Any = _as_int(raw) if name == "tax_year" else _clean_text(raw)
        if isinstance(value, str) and _REPLACEMENT_CHAR in value:
            # El portal sirve el archivo declarando UTF-8 con bytes en ISO-8859-1, asi
            # que algunos nombres llegan con caracteres irrecuperables.
            warnings.append(
                ReadingWarning(
                    code="TEXT_ENCODING_DAMAGED",
                    message=(
                        f"El portal entregó {label} con caracteres ilegibles. "
                        "Es un defecto conocido de la DIAN, no del documento."
                    ),
                    source=cell,
                )
            )
        fields.append(ExtractedField(name=name, value=value, source=cell))
    return fields


def _read_thresholds(sheet: Worksheet) -> list[ExtractedField]:
    """Los cinco topes de obligacion, tal como los calcula la DIAN."""
    fields: list[ExtractedField] = []
    for row, code in _THRESHOLD_ROWS.items():
        amount = _as_int(sheet.cell(row=row, column=_COL_AMOUNT).value)
        if amount is None:
            continue
        fields.append(
            ExtractedField(
                name=f"tope_{code.value}",
                value=amount,
                source=f"F{row}",
                unit="COP",
            )
        )
    return fields


def _read_details(sheet: Worksheet, warnings: list[ReadingWarning]) -> list[ExtractedRow]:
    """El detalle de lo que cada tercero reporto."""
    rows: list[ExtractedRow] = []
    for row in range(_FIRST_DETAIL_ROW, sheet.max_row + 1):
        amount = _as_int(sheet.cell(row=row, column=_COL_AMOUNT).value)
        concept = _clean_text(sheet.cell(row=row, column=_COL_CONCEPT).value)
        if amount is None or not concept:
            continue

        suggested = _clean_text(sheet.cell(row=row, column=_COL_SUGGESTED_USE).value) or ""
        rows.append(
            ExtractedRow(
                source=f"fila {row}",
                values={
                    "reporter_nit": _clean_text(
                        sheet.cell(row=row, column=_COL_REPORTER_NIT).value
                    ),
                    "reporter_name": _clean_text(
                        sheet.cell(row=row, column=_COL_REPORTER_NAME).value
                    ),
                    # A quien dice el tercero que le reporto. No siempre es el contribuyente.
                    "reported_id_number": _clean_text(
                        sheet.cell(row=row, column=_COL_REPORTED_ID).value
                    ),
                    "reported_name": _clean_text(
                        sheet.cell(row=row, column=_COL_REPORTED_NAME).value
                    ),
                    "concept": concept,
                    "concept_code": _concept_code(concept),
                    "amount": amount,
                    # La DIAN indica el renglon del 210 y el tope al que cuenta.
                    "form_lines": _form_lines(suggested),
                    "thresholds": _thresholds_of(suggested),
                    # La DIAN compara esta fila contra otra fuente y toma la mayor, en vez
                    # de sumarla: no es un sumando del tope.
                    "compared_not_added": bool(_COMPARED_NOT_ADDED_RE.search(suggested)),
                    "suggested_use": suggested or None,
                    "extra": _clean_text(sheet.cell(row=row, column=_COL_EXTRA).value),
                },
            )
        )

    if not rows:
        warnings.append(
            ReadingWarning(
                code="NO_REPORTED_ITEMS",
                message="el reporte no trae conceptos reportados por terceros",
            )
        )
    return rows


def _check_reported_to_taxpayer(
    fields: list[ExtractedField],
    rows: list[ExtractedRow],
    warnings: list[ReadingWarning],
) -> None:
    """Marca cada fila segun si el tercero se la reporto al titular, y avisa si no.

    Es la pregunta "y este ingreso que no es mio?" contestada con datos: el reporte dice a
    quien le reporto cada tercero, asi que se puede comparar con el titular. Importa porque
    esos valores entran a los topes de obligacion y a los renglones sugeridos como si fueran
    suyos, y quien responde por la declaracion es el titular.

    La conclusion queda en la propia fila (`reported_to_titular`) para que nadie mas tenga
    que volver a decidir cuando dos nombres son la misma persona, y ademas se avisa una vez
    por cada tercero, no una por fila: cuando un tercero confunde a una persona lo hace en
    todos los valores que le reporta, y cinco avisos del mismo error esconden los demas.
    """
    values = {f.name: f.value for f in fields}
    taxpayer_id = str(values.get("id_number") or "").strip()
    taxpayer_words = _name_words(values.get("taxpayer_name"))

    grouped: dict[tuple[str, str], list[ExtractedRow]] = {}
    for row in rows:
        reported_id = str(row.values.get("reported_id_number") or "").strip()
        reported_words = _name_words(row.values.get("reported_name"))

        if taxpayer_id and reported_id and reported_id != taxpayer_id:
            # Distinto numero de identificacion: el valor no es de este contribuyente.
            key = ("id", reported_id)
        elif (
            len(taxpayer_words) >= _MIN_SHARED_NAME_WORDS
            and len(reported_words) >= _MIN_SHARED_NAME_WORDS
            and len(taxpayer_words & reported_words) < _MIN_SHARED_NAME_WORDS
        ):
            # El numero es el del contribuyente pero el nombre es de otra persona.
            key = ("name", " ".join(sorted(reported_words)))
        else:
            # Coincide, o no hay con que comparar (nombre ausente o ilegible): en ninguno de
            # los dos casos se puede afirmar que sea de otra persona.
            row.values["reported_to_titular"] = True
            continue

        row.values["reported_to_titular"] = False
        grouped.setdefault(key, []).append(row)

    for (kind, _), affected in grouped.items():
        total = sum(int(r.values.get("amount") or 0) for r in affected)
        reporter = affected[0].values.get("reporter_name") or "un tercero"
        reported = affected[0].values.get("reported_name") or "otra persona"
        if kind == "id":
            message = (
                f"{reporter} reportó ${total:,.0f} a un número de identificación distinto "
                f"al del titular: ese valor no debería contar como suyo"
            ).replace(",", ".")
        else:
            message = (
                f"{reporter} reportó ${total:,.0f} al número de identificación del titular "
                f"pero a nombre de {reported}: hay que confirmar si ese valor es suyo"
            ).replace(",", ".")
        warnings.append(
            ReadingWarning(
                code="REPORTED_TO_ANOTHER_PERSON",
                message=message,
                source=affected[0].source,
            )
        )


def _name_words(name: Any) -> set[str]:
    """Palabras de un nombre, normalizadas para poder compararlo escrito de otra forma.

    Se quitan acentos y mayusculas, y se descartan las palabras de una sola letra (iniciales)
    y las que traigan el caracter de reemplazo que deja el portal cuando manda mal el
    encoding, porque no se puede saber que letra era.
    """
    if not name:
        return set()
    plain = unicodedata.normalize("NFKD", str(name))
    plain = "".join(c for c in plain if not unicodedata.combining(c)).upper()
    return {
        word
        for word in re.split(r"[^A-Z0-9]+", plain)
        if len(word) > 1 and _REPLACEMENT_CHAR not in word
    }


# ─────────────────────────── utilidades ───────────────────────────


def _concept_code(concept: str) -> str | None:
    """Extrae el codigo oficial del concepto del texto del detalle."""
    match = _CONCEPT_CODE_RE.search(concept)
    return match.group(1) if match else None


def _form_lines(suggested_use: str) -> list[int]:
    """Renglones del formulario 210 a los que la DIAN asigna el valor."""
    return sorted({int(n) for n in _FORM_LINE_RE.findall(suggested_use)})


def _thresholds_of(suggested_use: str) -> list[str]:
    """Topes de obligacion a los que el valor cuenta, por su codigo interno."""
    order = list(_THRESHOLD_ROWS.values())
    found: list[str] = []
    for number in _THRESHOLD_LABEL_RE.findall(suggested_use):
        index = int(number) - 1
        if 0 <= index < len(order):
            found.append(order[index].value)
    return sorted(set(found))


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        # El portal escribe fechas como datetime; se normalizan a ISO para que el valor
        # sea serializable y comparable.
        return value.date().isoformat() if value.time() == time.min else value.isoformat()
    text = str(value).strip()
    return re.sub(r"\s+", " ", text) or None


def _as_int(value: Any) -> int | None:
    """Convierte a entero los montos, que el portal escribe como texto."""
    if value is None:
        return None
    if isinstance(value, int | float):
        return int(value)
    digits = re.sub(r"[^\d-]", "", str(value))
    return int(digits) if digits and digits != "-" else None
