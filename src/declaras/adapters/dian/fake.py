"""Conector falso, deterministico, para desarrollo y pruebas.

Existe por dos razones concretas. Primera: el equipo que consume la API (el agente de
WhatsApp) puede integrarse hoy, sin esperar la calibracion del portal real. Segunda:
las ramas de error del Muisca son casi imposibles de provocar a voluntad, y aca se
disparan a pedido.

La rama se escoge con la clave enviada:
    contiene "bad"        -> DIAN_INVALID_CREDENTIALS
    contiene "locked"     -> DIAN_ACCOUNT_LOCKED
    contiene "down"       -> DIAN_PORTAL_UNAVAILABLE
    contiene "slow"       -> DIAN_PORTAL_TIMEOUT
    contiene "challenge"  -> reto de identidad (patron relevo); se resuelve con "1234"
    contiene "noexo"      -> exogena no publicada, el resto si baja
    contiene "sindecl"    -> primerizo: no hay declaracion anterior ni borrador
    cualquier otra        -> exito completo

EL ESCENARIO "sindecl" NO ES HIPOTETICO. Es lo que le pasa al primer contribuyente real del
producto, verificado contra el portal el 2026-08-08: la DIAN responde 404 a las dos consultas
de declaraciones y la extraccion entrega tres documentos de cinco. Faltaba como escenario, asi
que todo el camino que sigue —conciliacion, liquidacion y sobre todo las dos comparaciones, que
dependen justo de esos dos documentos— solo se ejercitaba con el caso feliz.
"""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from uuid import uuid4

from declaras.domain.errors import (
    DianAccountLockedError,
    DianDocumentUnavailableError,
    DianIdentityChallengeError,
    DianInvalidCredentialsError,
    DianPortalUnavailableError,
    DianSessionExpiredError,
    DianTimeoutError,
    ValidationError,
)
from declaras.domain.models import (
    BorradorEscrito,
    ChallengeAnswer,
    ChallengeKind,
    DianCredentials,
    DocumentType,
    IdentityChallenge,
    RawDocument,
    TaxpayerRef,
)

_EXPECTED_CHALLENGE_ANSWER = "1234"
_PDF_STUB = b"%PDF-1.4\n%% documento de prueba declaras\n"

# Los dos documentos que salen de `api.dian.gov.co` y que un primerizo no tiene.
_DECLARACIONES = frozenset({DocumentType.PRIOR_RETURN, DocumentType.SUGGESTED_RETURN})


def _exogena_sintetica(taxpayer: TaxpayerRef) -> bytes:
    """Un reporte de exogena con la MISMA disposicion que el real.

    Las posiciones no son decorativas: el lector busca la cedula en C7, el anio en C4 y los cinco
    topes en las filas 15 a 19. Una hoja con los mismos datos en otro sitio se lee vacia.

    Los topes van por ENCIMA del limite legal a proposito: el escenario util para probar es el
    de alguien obligado, que es donde el resultado tiene consecuencias.
    """
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws["C1"] = "Consulta de Información reportada por terceros"
    ws["H2"] = "2026-01-01 00:00:00"
    ws["C3"] = f"{taxpayer.tax_year}-12-31 00:00:00"
    ws["C4"] = str(taxpayer.tax_year)
    ws["C6"] = "C. C."
    ws["C7"] = taxpayer.id_number
    ws["C8"] = "PEREZ GOMEZ ANA MARIA"
    for columna, titulo in enumerate(
        [
            "NIT",
            "Nombre",
            "NIT",
            "Nombre reportado",
            "Detalle",
            "Valor",
            "Uso declaración Sugerida",
        ],
        start=1,
    ):
        ws.cell(row=14, column=columna, value=titulo)
    topes = [
        ("Tope 1 - Ingresos", 90_000_000),
        ("Tope 2 - Patrimonio", 30_000_000),
        ("Tope 3 - Consumo TC", 12_000_000),
        ("Tope 4 - Movimiento", 40_000_000),
        ("Tope 5 - Compras", 8_000_000),
    ]
    for offset, (etiqueta, valor) in enumerate(topes):
        ws.cell(row=15 + offset, column=5, value=etiqueta)
        ws.cell(row=15 + offset, column=6, value=valor)
    ws.cell(row=20, column=1, value="900111222")
    ws.cell(row=20, column=2, value="EMPRESA DEMO SAS")
    ws.cell(row=20, column=3, value=taxpayer.id_number)
    ws.cell(row=20, column=4, value="PEREZ GOMEZ ANA MARIA")
    ws.cell(row=20, column=5, value="Pagos por salarios (Concepto: 2276)")
    ws.cell(row=20, column=6, value=90_000_000)
    ws.cell(row=20, column=7, value="Tope 1: Ingresos brutos | R32 Ingresos brutos")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class FakeDianSession:
    """Sesion falsa que devuelve documentos sinteticos estables."""

    def __init__(self, *, scenario: str, challenge: IdentityChallenge | None) -> None:
        self.session_id = str(uuid4())
        self._scenario = scenario
        self._challenge = challenge
        self._closed = False

    @property
    def pending_challenge(self) -> IdentityChallenge | None:
        return self._challenge

    async def download(self, doc_type: DocumentType, taxpayer: TaxpayerRef) -> RawDocument:
        if self._challenge is not None:
            raise DianIdentityChallengeError(challenge=self._challenge.model_dump(mode="json"))
        if self._closed:
            raise ValidationError("La sesión ya fue cerrada.")
        if doc_type is DocumentType.EXOGENA and "noexo" in self._scenario:
            raise DianDocumentUnavailableError(
                "La DIAN todavía no publica la información exógena del periodo.",
                doc_label="exogena",
            )
        if doc_type in _DECLARACIONES and "sindecl" in self._scenario:
            # Se cita el `mensaje` real de la DIAN, igual que hace el conector HTTP, para que lo
            # que se prueba aca sea del mismo material que llega en produccion.
            raise DianDocumentUnavailableError(
                'La DIAN no reportó ninguna declaración. La DIAN respondió: "Documentos no '
                'encontrados". Si el contribuyente sí declaró, hay que verificarlo en el '
                "portal: puede que la consulta haya fallado y no que la declaración no exista.",
                doc_type=doc_type.value,
                evidencia="respuesta 404 de la API: Documentos no encontrados",
            )
        # LA EXOGENA ES UNA HOJA DE CALCULO, NO UN PDF. El falso devolvia el mismo relleno de
        # PDF para todo, asi que cualquier camino que la LEYERA de verdad —y no solo la
        # guardara— reventaba con "el archivo no es una hoja de calculo que se pueda leer".
        # Lo descubrio la consulta de obligacion, que es el primer camino que la parsea al vuelo.
        if doc_type is DocumentType.EXOGENA:
            return RawDocument(
                doc_type=doc_type,
                filename=f"exogena-{taxpayer.tax_year}.xlsx",
                content=_exogena_sintetica(taxpayer),
                content_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                source_url="https://fake.local/exogena",
                metadata={"fake": True, "tax_year": taxpayer.tax_year},
            )
        return RawDocument(
            doc_type=doc_type,
            filename=f"{doc_type.value.lower()}-{taxpayer.tax_year}.pdf",
            content=_PDF_STUB + f"{doc_type.value}:{taxpayer.subject_key}".encode(),
            content_type="application/pdf",
            source_url=f"https://fake.local/{doc_type.value.lower()}",
            metadata={"fake": True, "tax_year": taxpayer.tax_year},
        )

    async def listar_declaraciones(self) -> list[dict[str, object]]:
        """Un historial con un HUECO a proposito.

        El caso interesante no es la serie completa: es el año que falta en la mitad, porque
        eso es un atraso y el sistema tiene que saber mostrarlo. Aca falta 2023, con años
        declarados a ambos lados, que es la forma en que un hueco se puede AFIRMAR.

        EL LISTADO INCLUYE EL AÑO ANTERIOR (2024) porque `download` entrega su PRIOR_RETURN sin
        chistar: un falso que dijera no tenerlo mientras lo entrega estaria enseñando una
        realidad que no existe, y las pruebas que se apoyen en el probarian otra cosa.
        """
        if self._closed:
            raise ValidationError("La sesión ya fue cerrada.")
        if "sindecl" in self._scenario:
            return []
        return [
            {"anio": anio, "form_id": f"fake-{anio}"}
            for anio in (2025, 2024, 2022, 2021, 2020)
        ]

    async def descargar_declaracion(self, anio: int) -> RawDocument:
        if self._closed:
            raise ValidationError("La sesión ya fue cerrada.")
        disponibles = {d["anio"] for d in await self.listar_declaraciones()}
        if anio not in disponibles:
            raise DianDocumentUnavailableError(
                f"La DIAN no tiene la declaración presentada del año gravable {anio}.",
                doc_type=DocumentType.FILED_RETURN.value,
                tax_year=anio,
                available_years=sorted(disponibles, reverse=True),
            )
        return RawDocument(
            doc_type=DocumentType.FILED_RETURN,
            filename=f"declaracion-{anio}.pdf",
            content=_PDF_STUB + f"FILED_RETURN:{anio}".encode(),
            content_type="application/pdf",
            source_url=f"https://fake.local/declaracion/{anio}",
            metadata={"fake": True, "historial": True, "tax_year": anio},
        )

    async def escribir_borrador(
        self, taxpayer: TaxpayerRef, casillas: dict[int, int]
    ) -> BorradorEscrito:
        """Simula la escritura con los mismos desenlaces del portal real."""
        if self._closed:
            raise DianSessionExpiredError("La sesión ya fue cerrada.")
        # `sinborrador` YA NO FALLA: la cuenta sin borrador del año es el caso normal de un
        # primerizo, y el adaptador real lo crea copiando lo que hace el portal. El escenario
        # se conserva para fijar justamente eso — que no hay paso manual.
        # El id imita la forma real (los formularios del portal empiezan por 21) y el
        # resultado declara verificado: el fake no tiene un portal que corrompa nada.
        return BorradorEscrito(
            form_id="2118740000000",
            anio=taxpayer.tax_year,
            escritas=len(casillas),
            verificado=True,
            ajenas={},
        )

    async def capture_evidence(self, label: str) -> RawDocument:
        return RawDocument(
            doc_type=DocumentType.EVIDENCE,
            filename=f"{label}.png",
            content=b"\x89PNG\r\n\x1a\n fake",
            content_type="image/png",
            metadata={"label": label, "fake": True},
        )

    async def answer_challenge(self, answer: ChallengeAnswer) -> None:
        if self._challenge is None:
            raise ValidationError("La sesión no tiene ninguna verificación pendiente.")
        if answer.answers[0].strip() != _EXPECTED_CHALLENGE_ANSWER:
            raise DianInvalidCredentialsError("La respuesta de verificación fue rechazada.")
        self._challenge = None

    async def close(self) -> None:
        self._closed = True


class FakeDianConnector:
    """Implementa DianConnector sin tocar la red."""

    def __init__(self, *, challenge_ttl_s: int = 600) -> None:
        self._challenge_ttl_s = challenge_ttl_s

    async def open_session(
        self, credentials: DianCredentials, taxpayer: TaxpayerRef
    ) -> FakeDianSession:
        scenario = credentials.password.get_secret_value().lower()

        if "bad" in scenario:
            raise DianInvalidCredentialsError()
        if "locked" in scenario:
            raise DianAccountLockedError()
        if "down" in scenario:
            raise DianPortalUnavailableError()
        if "slow" in scenario:
            raise DianTimeoutError()

        challenge = None
        if "challenge" in scenario:
            now = datetime.now(UTC)
            challenge = IdentityChallenge(
                kind=ChallengeKind.EMAIL_CODE,
                prompt="Ingresa el codigo de 4 digitos que llego a tu correo",
                options=[],
                issued_at=now,
                expires_at=now + timedelta(seconds=self._challenge_ttl_s),
            )
        return FakeDianSession(scenario=scenario, challenge=challenge)

    async def shutdown(self) -> None:
        return None
