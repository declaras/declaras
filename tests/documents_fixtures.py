"""Constructores de documentos sinteticos para pruebas.

Los archivos reales de un contribuyente no se versionan (traen datos personales y
financieros reales). Estos constructores reproducen la MISMA estructura que valida cada
parser, calibrada contra un documento real, pero con datos inventados.
"""

from __future__ import annotations

from io import BytesIO

from fpdf import FPDF
from openpyxl import Workbook


def build_exogena_xlsx(
    *,
    tax_year: int = 2025,
    id_number: str = "1000000001",
    taxpayer_name: str = "PEREZ GOMEZ ANA MARIA",
    thresholds: dict[str, int] | None = None,
    detail_rows: list[dict] | None = None,
) -> bytes:
    """Reproduce el reporte de informacion exogena, calibrado contra el XLSX real."""
    thresholds = thresholds or {
        "ingresos": 10_000_000,
        "patrimonio": 5_000_000,
        "consumo_tarjeta": 1_000_000,
        "movimientos": 20_000_000,
        "compras": 500_000,
    }
    detail_rows = (
        detail_rows
        if detail_rows is not None
        else [
            {
                "reporter_nit": "900111222",
                "reporter_name": "EMPRESA DEMO SAS",
                "concept": "Valor ingreso laboral promedio (Concepto: 2276)",
                "amount": 3_000_000,
                "suggested_use": "R36 Otras rentas exentas",
            },
            {
                "reporter_nit": "900333444",
                "reporter_name": "BANCO DEMO",
                "concept": "Retencion en la fuente (Concepto: 5004)",
                "amount": 150_000,
                "suggested_use": "R132 Retenciones año gravable a declarar",
            },
        ]
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws["C1"] = "Consulta de Información reportada por terceros"
    ws["G2"] = "Fecha   Reporte:"
    ws["H2"] = "2026-01-01 00:00:00"
    ws["A3"] = "Fecha corte del proceso: "
    ws["C3"] = "2025-12-31 00:00:00"
    ws["A4"] = "Año al que se refiere la consulta:"
    ws["C4"] = str(tax_year)
    ws["A6"] = "Tipo de documento:"
    ws["C6"] = "C. C."
    ws["A7"] = "Identificación:"
    ws["C7"] = id_number
    ws["A8"] = "Nombres / Razón social:"
    ws["C8"] = taxpayer_name

    ws["A14"] = "NIT"
    ws["B14"] = "Nombre / Razón Social"
    ws["C14"] = "NIT"
    ws["D14"] = "Nombre/Razón Social reportada por el tercero"
    ws["E14"] = "Detalle"
    ws["F14"] = "Valor"
    ws["G14"] = "Uso declaración Sugerida"

    threshold_order = ["ingresos", "patrimonio", "consumo_tarjeta", "movimientos", "compras"]
    labels = [
        "Tope 1 - Ingresos",
        "Tope 2 - Patrimonio",
        "Tope 3 - Consumo TC",
        "Tope 4 - Movimiento",
        "Tope 5 - Compras",
    ]
    for offset, (code, label) in enumerate(zip(threshold_order, labels, strict=True)):
        row = 15 + offset
        ws.cell(row=row, column=5, value=label)
        ws.cell(row=row, column=6, value=thresholds[code])

    for offset, item in enumerate(detail_rows):
        row = 20 + offset
        ws.cell(row=row, column=1, value=item["reporter_nit"])
        ws.cell(row=row, column=2, value=item["reporter_name"])
        # A quien dice el tercero que le reporto. Por defecto es el titular, pero un caso
        # puede pasar otro nombre o otra identificacion para reproducir un reporte cruzado.
        ws.cell(row=row, column=3, value=item.get("reported_id_number", id_number))
        ws.cell(row=row, column=4, value=item.get("reported_name", taxpayer_name))
        ws.cell(row=row, column=5, value=item["concept"])
        ws.cell(row=row, column=6, value=item["amount"])
        ws.cell(row=row, column=7, value=item["suggested_use"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_einvoice_summary_xlsx(
    *,
    tax_year: int = 2025,
    id_number: str = "1000000001",
    taxpayer_name: str = "PEREZ GOMEZ ANA MARIA",
    invoices: list[dict] | None = None,
) -> bytes:
    """Reproduce el resumen de facturas electronicas, calibrado contra el XLSX real."""
    invoices = (
        invoices
        if invoices is not None
        else [
            {
                "issuer_nit": "900555666",
                "issuer_name": "SUPERMERCADO DEMO SAS",
                "issue_date": "2025-03-15",
                "invoiced_amount": 120_000,
                "net_amount": 120_000,
                "benefit_eligible_amount": 120_000,
                "payment_method": "Electrónicos",
                "invoice_number": "F001",
                "cufe": "abc123",
            },
            {
                "issuer_nit": "900777888",
                "issuer_name": "TIENDA DEMO",
                "issue_date": "2025-05-20",
                "invoiced_amount": 50_000,
                "net_amount": 50_000,
                "benefit_eligible_amount": 0,
                "payment_method": "Efectivo",
                "invoice_number": "F002",
                "cufe": "def456",
            },
        ]
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"], ws["B1"] = "Nombre", "Informe de Facturas electrónicas adquiridas"
    ws["A3"], ws["B3"] = "Año Gravable", str(tax_year)
    ws["A4"], ws["B4"] = "Tipo Documento Adquiriente", "13 Cédula de Ciudadanía"
    ws["A5"], ws["B5"] = "Doc Identificacion Adq.", id_number
    ws["A6"], ws["B6"] = "NIT", id_number
    ws["A7"], ws["B7"] = "Nombre o razón social", taxpayer_name

    headers = [
        "Identificación Emisor Factura",
        "Nombre Emisor Factura",
        "Fecha Emisión",
        "Valor Facturado",
        "Valor Notas Crédito",
        "Valor Notas Débito",
        "Valor Factura / Afectada con Notas",
        "Valor Susceptible Beneficio",
        "Medios De Pago",
        "Num_factura_venta",
        "CUFE",
    ]
    for col, label in enumerate(headers, start=1):
        ws.cell(row=25, column=col, value=label)

    for offset, inv in enumerate(invoices):
        row = 26 + offset
        ws.cell(row=row, column=1, value=inv["issuer_nit"])
        ws.cell(row=row, column=2, value=inv["issuer_name"])
        ws.cell(row=row, column=3, value=inv["issue_date"])
        ws.cell(row=row, column=4, value=inv["invoiced_amount"])
        ws.cell(row=row, column=5, value=0)
        ws.cell(row=row, column=6, value=0)
        ws.cell(row=row, column=7, value=inv["net_amount"])
        ws.cell(row=row, column=8, value=inv["benefit_eligible_amount"])
        ws.cell(row=row, column=9, value=inv["payment_method"])
        ws.cell(row=row, column=10, value=inv["invoice_number"])
        ws.cell(row=row, column=11, value=inv["cufe"])

    footer_row = 26 + len(invoices) + 1
    ws.cell(row=footer_row, column=1, value=f"{len(invoices)} Facturas procesadas disponibles")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_rut_pdf(
    *,
    form_number: str = "999888777666",
    nit: str = "1000000001",
    verification_digit: str = "5",
    collection_office: str = "Impuestos de Medellin",
    taxpayer_kind: str = "Persona natural o sucesión ilíquida",
    id_kind: str = "Cédula de Ciudadanía",
    id_number: str = "1000000001",
    last_name_1: str = "PEREZ",
    last_name_2: str = "GOMEZ",
    first_name_1: str = "ANA",
    other_names: str = "MARIA",
    email: str = "ana.perez@example.com",
    economic_activity_code: str = "6201",
    activity_start_date: str = "20200115",
) -> bytes:
    """Reproduce el RUT en PDF: un bloque de plantilla y luego los valores en orden
    secuencial, tal como lo dibuja el generador real del portal (calibrado el
    2026-07-25 contra un RUT real)."""

    def spaced(digits: str) -> str:
        return " ".join(digits)

    template_blob = "IDENTIFICACIÓN " + ("Formulario del Registro Único Tributario " * 15)
    value_tokens = [
        "Actualización de oficio",
        "1 3",
        form_number,
        spaced(nit),
        f" {verification_digit}",
        f" {collection_office}",
        " 3 2",
        taxpayer_kind,
        " 2",
        f" {id_kind}",
        " 1 3",
        f" {spaced(id_number)}",
        "COLOMBIA",
        " 1 6 9",
        " Antioquia",
        " 0 5",
        "Medellín",
        " 0 0 1",
        last_name_1,
        f" {last_name_2}",
        f" {first_name_1}",
        f" {other_names}",
        "COLOMBIA",
        " 1 6 9",
        " Antioquia",
        " 0 5",
        "Medellín",
        " 0 0 1",
        "CL 10 # 20 - 30",
        email,
        "   3001234567",
        spaced(economic_activity_code),
        spaced(activity_start_date),
    ]

    pdf = FPDF()
    pdf.set_auto_page_break(False)
    pdf.add_page()
    pdf.set_font("Helvetica", size=7)
    # Una sola llamada de texto: en el RUT real la plantilla se dibuja como UN fragmento
    # contiguo, no envuelto en varias lineas. Reproducirlo asi es lo que hace que el
    # detector de "fin de la plantilla" se comporte igual que con el documento real.
    pdf.text(10, 15, template_blob)

    y = 120
    for token in value_tokens:
        pdf.set_xy(10, y)
        pdf.cell(0, 4, token)
        y += 4.2

    return bytes(pdf.output())


def build_renta_210_pdf(
    *,
    tax_year: int = 2024,
    id_number: int = 1020304050,
    verification_digit: int = 7,
    patrimonio_bruto: int = 50_000_000,
    deudas: int = 20_000_000,
    patrimonio_liquido: int = 30_000_000,
    ingresos_brutos: int = 80_000_000,
    no_constitutivos: int = 6_000_000,
    renta_liquida: int = 74_000_000,
    ingresos_no_laborales: int = 5_000_000,
) -> bytes:
    """Reproduce el formulario 210 con la geometria real del PDF del portal.

    Se escribe el flujo de contenido a mano en vez de usar `fpdf` porque lo que el parser
    lee son las coordenadas de cada valor, y `fpdf` no permite controlarlas al nivel que
    hace falta: solo conserva la matriz del primer texto de la pagina.

    Las coordenadas son las de una declaracion real del ano gravable 2024, incluida la parte
    que mas facil se rompe: la franja del patrimonio y la de totales no usan las mismas
    columnas que la cedula general.
    """
    trazos: list[tuple[float, float, str]] = []

    def poner(x: float, y: float, texto: str) -> None:
        trazos.append((x, y, texto))

    def miles(valor: int) -> str:
        return f"{valor:,}"

    # Cabecera: el ano gravable y la cedula se imprimen digito por digito. El digito de
    # verificacion va separado por un espacio mayor.
    for i, digito in enumerate(str(tax_year)):
        poner(67.9 + i * 10.75, 722.5, digito)
    for i, digito in enumerate(str(id_number)):
        poner(86.7 + i * 10.9, 615.5, digito)
    poner(199.5 + (len(str(id_number)) - 10) * 10.9, 615.5, str(verification_digit))

    # Patrimonio: tres casillas con su propio juego de columnas.
    poner(216.4, 593.5, miles(patrimonio_bruto))
    poner(366.5, 593.5, miles(deudas))
    poner(585.5, 593.5, miles(patrimonio_liquido))

    # Cedula general: cuatro columnas. La primera lleva las rentas de trabajo y la cuarta
    # las rentas no laborales.
    cedula = [
        (569.5, ingresos_brutos, ingresos_no_laborales),
        (545.5, no_constitutivos, 0),
        (521.5, renta_liquida, ingresos_no_laborales),
    ]
    for y, trabajo, no_laboral in cedula:
        poner(190.0, y, miles(trabajo))
        poner(345.6, y, "0")
        poner(465.6, y, "0")
        poner(554.4, y, miles(no_laboral))

    # Fila que el formulario solo abre en la ultima columna (devoluciones y descuentos).
    poner(585.5, 557.5, "0")

    # Renta liquida ordinaria y franja de totales.
    poner(190.0, 377.5, miles(renta_liquida))
    poner(554.4, 377.5, miles(ingresos_no_laborales))
    total_cedula = renta_liquida + ingresos_no_laborales
    poner(130.0, 365.5, miles(total_cedula))
    poner(272.4, 365.5, "0")
    poner(406.0, 365.5, miles(total_cedula))
    poner(585.5, 365.5, "0")

    return _build_pdf_with_positioned_text(trazos)


def _build_pdf_with_positioned_text(trazos: list[tuple[float, float, str]]) -> bytes:
    """Arma un PDF minimo donde cada texto lleva su propia matriz de posicion."""
    operaciones = "".join(
        f"BT /F1 8 Tf 1 0 0 1 {x} {y} Tm ({texto})Tj ET\n" for x, y, texto in trazos
    )
    objetos = [
        b"<</Type/Catalog/Pages 2 0 R>>",
        b"<</Type/Pages/Kids[3 0 R]/Count 1>>",
        b"<</Type/Page/Parent 2 0 R/MediaBox[0 0 595.28 841.89]"
        b"/Resources<</Font<</F1 4 0 R>>>>/Contents 5 0 R>>",
        b"<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>",
    ]
    flujo = operaciones.encode("latin-1")
    objetos.append(b"<</Length %d>>stream\n%s\nendstream" % (len(flujo), flujo))

    salida = bytearray(b"%PDF-1.4\n")
    posiciones: list[int] = []
    for numero, cuerpo in enumerate(objetos, start=1):
        posiciones.append(len(salida))
        salida += b"%d 0 obj\n" % numero + cuerpo + b"\nendobj\n"
    inicio_xref = len(salida)
    salida += b"xref\n0 %d\n0000000000 65535 f \n" % (len(objetos) + 1)
    for posicion in posiciones:
        salida += b"%010d 00000 n \n" % posicion
    salida += b"trailer\n<</Size %d/Root 1 0 R>>\nstartxref\n%d\n%%%%EOF\n" % (
        len(objetos) + 1,
        inicio_xref,
    )
    return bytes(salida)
